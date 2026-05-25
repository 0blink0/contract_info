"""Export dictionary sheets from 产品要素-2.xlsx to dicts/*.json."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = PROJECT_ROOT / "example" / "产品要素-2.xlsx"
DICTS_DIR = PROJECT_ROOT / "dicts"

TARGET_SHEETS = [
    "销售机构及名称对应表",
    "业绩比较基准指数表",
    "产品分类字典",
]


def slugify(name: str) -> str:
    s = re.sub(r"[^\w\u4e00-\u9fff]+", "_", name.strip())
    return s.strip("_") or "sheet"


def sheet_to_records(ws) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(row):
            continue
        item: dict[str, str] = {}
        for i, col in enumerate(header):
            if not col:
                continue
            val = row[i] if i < len(row) else None
            if val is None:
                continue
            item[col] = str(val).strip()
        if item:
            records.append(item)
    return records


def main() -> None:
    if not XLSX_PATH.is_file():
        raise SystemExit(f"Missing workbook: {XLSX_PATH}")

    DICTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    exported = 0
    for sheet_name in wb.sheetnames:
        if sheet_name not in TARGET_SHEETS and "字典" not in sheet_name and "对应表" not in sheet_name:
            continue
        ws = wb[sheet_name]
        records = sheet_to_records(ws)
        if not records:
            continue
        out_path = DICTS_DIR / f"{slugify(sheet_name)}.json"
        out_path.write_text(
            json.dumps(records, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Wrote {out_path} ({len(records)} rows)")
        exported += 1

    wb.close()
    if exported == 0:
        raise SystemExit("No dictionary sheets exported")
    print(f"Done: {exported} file(s)")


if __name__ == "__main__":
    main()
