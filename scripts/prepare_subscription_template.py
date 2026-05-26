"""Copy example 申赎母版 to templates/ and fix sheet name."""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "example" / "产品申赎费率导入模板.xlsx"
DST = ROOT / "templates" / "产品申赎费率导入模板.xlsx"
SHEET = "产品申赎费率导入模板"


def main() -> int:
    if not SRC.is_file():
        print(f"Missing source: {SRC}", file=sys.stderr)
        return 1
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST)
    wb = load_workbook(DST)
    wb.active.title = SHEET
    wb.save(DST)
    wb.close()
    wb2 = load_workbook(DST, read_only=True)
    assert SHEET in wb2.sheetnames, wb2.sheetnames
    wb2.close()
    print(f"Wrote {DST}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
