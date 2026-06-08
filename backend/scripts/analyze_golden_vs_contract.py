"""Analyze golden xlsx vs contract extractability. Output UTF-8 report."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
EXAMPLE = ROOT / "example"

FUNDS = {
    "石云中证1000": {
        "docx": "石云中证1000资产进取一号私募证券投资基金-基金合同(1).docx",
        "code_golden": "HXZC38",  # from prior grep
    },
    "石云福禄1000": {
        "docx": "石云福禄1000指数增强一号私募证券投资基金(1).docx",
        "code_golden": "FLZB38",
    },
}

MANUAL_ONLY = {
    "销售经理",
    "产品经理",
    "投资机构",
    "补充协议",
    "特殊的产品类型",
    "运行状态",
}

SYSTEM_NOT_IN_CONTRACT = {
    "基金代码",  # 运营/系统清单
    "备案编码",  # 成立后才知，合同可能有
}

CRM_OR_POST_ESTABLISH = {
    "成立日期",
    "备案日期",
    "到期日期",
    "清算完成日",
    "清算起始日",
    "投资起始日",  # 有时合同有
}


def product_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["产品要素模板"]
    headers = [
        str(c).strip() if c else ""
        for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    ]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        if not row or not row[1]:
            continue
        d = {
            h: (str(v).strip() if v is not None else "")
            for h, v in zip(headers, row)
            if h
        }
        rows.append(d)
    wb.close()
    return rows


def fee_ops_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["产品运营费率导入模板"]
    headers = [
        str(c).replace("【必填】", "").replace("【非必填】", "").strip()
        if c
        else ""
        for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    ]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=50, values_only=True):
        if not row or not row[0]:
            continue
        d = {
            h: (str(v).strip() if v is not None else "")
            for h, v in zip(headers, row)
            if h
        }
        if d.get("基金名称") or d.get("运营费类型"):
            rows.append(d)
    wb.close()
    return rows


def sub_red_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [
        str(c).replace("【必填】", "").replace("【非必填】", "").strip()
        if c
        else ""
        for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    ]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=80, values_only=True):
        if not row or not row[0]:
            continue
        d = {
            h: (str(v).strip() if v is not None else "")
            for h, v in zip(headers, row)
            if h
        }
        rows.append(d)
    wb.close()
    return rows


def classify_product_field(field: str, has_value_in_golden: bool) -> str:
    if field in MANUAL_ONLY:
        return "manual_only"
    if field in SYSTEM_NOT_IN_CONTRACT:
        return "system_or_post" if field == "基金代码" else "contract_maybe"
    if field in CRM_OR_POST_ESTABLISH:
        return "contract_maybe"
    if field in ("投资顾问",) and not has_value_in_golden:
        return "contract_optional"
    return "contract_expected"


def main() -> None:
    prod_path = EXAMPLE / "产品要素 - 副本(1).xlsx"
    fee_path = EXAMPLE / "产品运营费率导入模板.xlsx"
    sub_path = EXAMPLE / "产品申赎费率导入模板.xlsx"

    products = product_rows(prod_path)
    fees = fee_ops_rows(fee_path)
    subs = sub_red_rows(sub_path)

    report: dict = {
        "product_by_fund": [],
        "fee_ops_summary": {},
        "sub_red_summary": {},
        "field_classification": {},
    }

    all_headers = set()
    for p in products:
        all_headers.update(p.keys())

    for p in products:
        name = p.get("基金全称", "")
        filled = {k: v for k, v in p.items() if v}
        empty = [k for k in p if not p.get(k)]
        by_cat = {"manual_only": [], "system_or_post": [], "contract_maybe": [], "contract_expected": [], "contract_optional": []}
        for k in all_headers:
            cat = classify_product_field(k, bool(p.get(k)))
            if not p.get(k):
                by_cat.setdefault("empty_" + cat, []).append(k)
            else:
                by_cat.setdefault("filled_" + cat, []).append(k)
        report["product_by_fund"].append(
            {
                "基金全称": name,
                "filled_count": len(filled),
                "empty_count": len(empty),
                "filled_fields": sorted(filled.keys()),
                "empty_fields": sorted(empty),
                "sample_values": {k: filled[k][:80] for k in list(filled.keys())[:12]},
            }
        )

    # classify all product columns
    for h in sorted(all_headers):
        if not h or h.startswith("Unnamed"):
            continue
        report["field_classification"][h] = classify_product_field(h, False)

    report["fee_ops_summary"] = {
        "row_count": len(fees),
        "types": sorted({r.get("运营费类型", "") for r in fees}),
        "funds": sorted({r.get("基金名称", "")[:30] for r in fees}),
    }
    report["sub_red_summary"] = {
        "row_count": len(subs),
        "fee_types": sorted({r.get("费用类型", "") for r in subs}),
        "sample_rates": [
            {r.get("基金名称", "")[:20]: r.get("费率", "")}
            for r in subs[:12]
            if r.get("费率")
        ],
    }

    out = ROOT / "example" / "_golden_field_analysis.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", out)


if __name__ == "__main__":
    main()
