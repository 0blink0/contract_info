from __future__ import annotations

import json
import shutil
import tempfile
import threading
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from copy import copy as _copy_obj

_index_lock = threading.Lock()

from openpyxl import load_workbook

from backend.app.config import data_dir, templates_dir
from backend.app.export.column_map import (
    FEE_DATA_START_ROW,
    FEE_HEADER_ROW,
    FEE_SHEET,
    LOCK_DATA_START_ROW,
    LOCK_HEADER_ROW,
    LOCK_SHEET,
    PRODUCT_DATA_ROW,
    PRODUCT_HEADER_ROW,
    PRODUCT_SHEET,
    SHARE_DATA_START_ROW,
    SHARE_HEADER_ROW,
    SHARE_SHEET,
    SUBSCRIPTION_DATA_START_ROW,
    SUBSCRIPTION_HEADER_ROW,
    SUBSCRIPTION_SHEET,
)
from backend.app.export.xlsx_utils import (
    build_header_index,
    clear_data_rows,
    keep_only_sheet,
    write_cell_values,
)

# Template filename (under templates_dir()) + target sheet name for each table type.
# Must stay in sync with backend/app/export/pipeline.py constants.
_TEMPLATE_FILES: dict[str, tuple[str, str]] = {
    "product-elements":      ("产品要素 - 副本(1).xlsx",        PRODUCT_SHEET),
    "fee-rates":             ("产品运营费率导入模板-1.xlsx",    FEE_SHEET),
    "lock-periods":          ("产品要素 - 副本(1).xlsx",        LOCK_SHEET),
    "share-classes":         ("产品要素 - 副本(1).xlsx",        SHARE_SHEET),
    "subscription-fee-rates":("产品申赎费率导入模板.xlsx",      SUBSCRIPTION_SHEET),
}

# (sheet_name, header_row_1indexed, data_start_row_1indexed)
_TABLE_READ_CONFIG: dict[str, tuple[str, int, int]] = {
    "product-elements": (PRODUCT_SHEET, PRODUCT_HEADER_ROW, PRODUCT_DATA_ROW),
    "fee-rates": (FEE_SHEET, FEE_HEADER_ROW, FEE_DATA_START_ROW),
    "lock-periods": (LOCK_SHEET, LOCK_HEADER_ROW, LOCK_DATA_START_ROW),
    "share-classes": (SHARE_SHEET, SHARE_HEADER_ROW, SHARE_DATA_START_ROW),
    "subscription-fee-rates": (SUBSCRIPTION_SHEET, SUBSCRIPTION_HEADER_ROW, SUBSCRIPTION_DATA_START_ROW),
}

TABLE_TYPE_LABELS: dict[str, str] = {
    "product-elements": "产品要素",
    "fee-rates": "运营费率",
    "lock-periods": "份额锁定期",
    "share-classes": "分级份额",
    "subscription-fee-rates": "申赎费率",
}


def _merges_dir() -> Path:
    p = data_dir() / "merges"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _index_path() -> Path:
    return _merges_dir() / "index.json"


def _load_index() -> list[dict]:
    p = _index_path()
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save_index(records: list[dict]) -> None:
    _index_path().write_text(
        json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _append_record(record: dict) -> None:
    """Thread-safe append of a single record to the index."""
    with _index_lock:
        records = _load_index()
        records.insert(0, record)
        _save_index(records)


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """Copy cell formatting from src_row to dst_row (values untouched)."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = _copy_obj(src.font)
            dst.border = _copy_obj(src.border)
            dst.fill = _copy_obj(src.fill)
            dst.number_format = src.number_format
            dst.protection = _copy_obj(src.protection)
            dst.alignment = _copy_obj(src.alignment)


def _read_xlsx_rows(
    path: Path,
    table_type: str,
) -> tuple[list[str], list[dict[str, str]]]:
    """Read non-empty-header columns and data rows from an exported xlsx file."""
    config = _TABLE_READ_CONFIG.get(table_type)
    sheet_name, header_row_idx, data_start_idx = config if config else (None, 1, 2)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < header_row_idx:
        return [], []

    raw_headers = all_rows[header_row_idx - 1]
    # Keep only columns with non-empty header text
    valid_cols: list[tuple[int, str]] = []
    for i, h in enumerate(raw_headers):
        label = str(h).strip() if h is not None else ""
        if label:
            valid_cols.append((i, label))

    if not valid_cols:
        return [], []

    headers = [label for _, label in valid_cols]
    data: list[dict[str, str]] = []

    for row in all_rows[data_start_idx - 1:]:
        padded = list(row) + [None] * max(0, len(raw_headers) - len(row))
        row_dict: dict[str, str] = {}
        for col_idx, label in valid_cols:
            v = padded[col_idx] if col_idx < len(padded) else None
            row_dict[label] = str(v).strip() if v is not None else ""
        if any(v for v in row_dict.values()):
            data.append(row_dict)

    return headers, data


def merge_xlsx_files(
    paths: list[Path],
    table_type: str,
) -> tuple[bytes, list[str], list[dict[str, str]]]:
    """Merge multiple xlsx files into one, using the original clean template as base.

    Uses the same template files as the export pipeline so the output has
    identical title rows, header formatting, and alternating row colors.
    Returns (xlsx_bytes, column_headers, preview_rows capped at 200).
    """
    read_cfg = _TABLE_READ_CONFIG.get(table_type)
    _, header_row_idx, data_start_idx = read_cfg if read_cfg else (None, 1, 2)

    # Collect data rows from all source files
    base_headers: list[str] = []
    combined_rows: list[dict[str, str]] = []
    for path in paths:
        headers, rows = _read_xlsx_rows(path, table_type)
        if not base_headers and headers:
            base_headers = headers
        combined_rows.extend(rows)

    # Resolve the clean template file (same as export pipeline uses)
    tpl_file_cfg = _TEMPLATE_FILES.get(table_type)
    tpl_path: Path | None = None
    target_sheet: str | None = None
    if tpl_file_cfg:
        tpl_filename, target_sheet = tpl_file_cfg
        candidate = templates_dir() / tpl_filename
        if candidate.is_file():
            tpl_path = candidate

    # Fall back to first source file if template not found
    base_src = tpl_path if tpl_path else paths[0]

    tmp_path = Path(tempfile.mktemp(suffix=".xlsx"))
    try:
        shutil.copy2(base_src, tmp_path)
        wb = load_workbook(tmp_path)

        # Keep only the target sheet (removes auxiliary sheets from the template)
        if target_sheet and target_sheet in wb.sheetnames:
            keep_only_sheet(wb, target_sheet)
            ws = wb[target_sheet]
        else:
            ws = wb.active

        # clear_data_rows removes values but preserves cell formatting/colors
        original_max_row = ws.max_row or data_start_idx
        max_col = ws.max_column or 1

        clear_data_rows(ws, data_start_idx)
        header_index = build_header_index(ws, header_row_idx)

        # The template's data rows may carry irregular colors (rows were grouped
        # by contract, not strictly alternating). Always re-apply the stripe by
        # copying from the first two data rows of the clean template.
        # Some dropdown columns (e.g. 申赎费类型, 计费方式, 费率类型) have the
        # same fill in both row A and row B of the template, so _copy_row_style
        # alone would never alternate them. Fix: extract the two reference fills
        # from a reliably-alternating column (col 1) and force-apply the correct
        # fill to every cell after the row-style copy.
        pattern_row_a = data_start_idx
        pattern_row_b = data_start_idx + 1 if original_max_row >= data_start_idx + 1 else data_start_idx
        fill_even = _copy_obj(ws.cell(row=pattern_row_a, column=1).fill)
        fill_odd  = _copy_obj(ws.cell(row=pattern_row_b, column=1).fill)

        for i, row_dict in enumerate(combined_rows):
            dst_row = data_start_idx + i
            pattern_src = pattern_row_a if i % 2 == 0 else pattern_row_b
            _copy_row_style(ws, pattern_src, dst_row, max_col)
            correct_fill = fill_even if i % 2 == 0 else fill_odd
            for col in range(1, max_col + 1):
                ws.cell(row=dst_row, column=col).fill = _copy_obj(correct_fill)
            write_cell_values(ws, dst_row, header_index, row_dict)

        # Delete leftover template rows after the last data row so stale
        # colors and values don't appear below the merged data.
        last_data_row = data_start_idx + len(combined_rows) - 1
        if ws.max_row > last_data_row:
            ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

        buf = BytesIO()
        wb.save(buf)
        wb.close()
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except PermissionError:
            pass

    return buf.getvalue(), base_headers, combined_rows[:200]


def create_merge(
    job_ids: list[str],
    table_type: str,
    source_jobs: list[dict],
    xlsx_bytes: bytes,
    columns: list[str],
    row_count: int,
    name: str = "",
) -> dict:
    merge_id = str(uuid.uuid4())
    file_path = _merges_dir() / f"{merge_id}.xlsx"
    file_path.write_bytes(xlsx_bytes)

    rel_path = str(file_path.relative_to(data_dir()))
    record: dict = {
        "id": merge_id,
        "name": name or f"{TABLE_TYPE_LABELS.get(table_type, table_type)} 合并",
        "table_type": table_type,
        "source_jobs": source_jobs,
        "merged_at": datetime.now(timezone.utc).isoformat(),
        "row_count": row_count,
        "file_path": rel_path,
        "columns": columns,
    }

    _append_record(record)
    return record


def list_merges() -> list[dict]:
    return _load_index()


def get_merge(merge_id: str) -> dict | None:
    for r in _load_index():
        if r.get("id") == merge_id:
            return r
    return None


def delete_merge(merge_id: str) -> bool:
    with _index_lock:
        records = _load_index()
        remaining = [r for r in records if r.get("id") != merge_id]
        if len(remaining) == len(records):
            return False
        for r in records:
            if r.get("id") == merge_id:
                fp = data_dir() / r.get("file_path", "")
                try:
                    fp.unlink(missing_ok=True)
                except Exception:
                    pass
        _save_index(remaining)
    return True


def delete_all_merges() -> int:
    with _index_lock:
        records = _load_index()
        for r in records:
            fp = data_dir() / r.get("file_path", "")
            try:
                fp.unlink(missing_ok=True)
            except Exception:
                pass
        _save_index([])
    return len(records)


def get_merge_file_path(merge_id: str) -> Path | None:
    record = get_merge(merge_id)
    if not record:
        return None
    fp = data_dir() / record.get("file_path", "")
    return fp if fp.is_file() else None


def append_to_merge(
    merge_id: str,
    new_job_paths: list[Path],
    new_source_jobs: list[dict],
) -> dict | None:
    """Append new job xlsx files to an existing merge record.

    Reads data from the existing merged file + new job files, re-writes the file,
    and updates the index record in place.  Returns the updated record or None if
    the merge_id is not found.
    """
    with _index_lock:
        records = _load_index()
        existing = next((r for r in records if r.get("id") == merge_id), None)
        if not existing:
            return None

        table_type: str = existing.get("table_type", "")
        existing_fp = data_dir() / existing.get("file_path", "")

        # Combine existing merged data with new job files
        all_paths: list[Path] = []
        if existing_fp.is_file():
            all_paths.append(existing_fp)
        all_paths.extend(new_job_paths)

        xlsx_bytes, columns, preview_rows = merge_xlsx_files(all_paths, table_type)

        # Overwrite the file in place
        existing_fp.write_bytes(xlsx_bytes)

        existing["source_jobs"] = (existing.get("source_jobs") or []) + new_source_jobs
        existing["merged_at"] = datetime.now(timezone.utc).isoformat()
        existing["row_count"] = len(preview_rows)
        existing["columns"] = columns

        _save_index(records)
        return existing


def get_merge_preview(merge_id: str, max_rows: int = 100) -> dict | None:
    """Return preview data by re-reading the merged xlsx with the correct header row."""
    record = get_merge(merge_id)
    if not record:
        return None
    fp = data_dir() / record.get("file_path", "")
    if not fp.is_file():
        return None

    table_type: str = record.get("table_type", "")
    headers, rows = _read_xlsx_rows(fp, table_type)
    return {"id": merge_id, "columns": headers, "rows": rows[:max_rows]}
