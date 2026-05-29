from __future__ import annotations



import uuid

from pathlib import Path

from typing import Any



from openpyxl import load_workbook



from backend.app.config import data_dir

SNIPPET_DISPLAY = "摘录原文"
_SKIP_PREVIEW_KEYS = frozenset({"snippet"})


def _append_snippet_column(
    columns_order: list[str],
    rows: list[dict[str, str | None]],
    raw_rows: list[dict[str, Any]],
) -> tuple[list[str], list[dict[str, str | None]]]:
    has_snippet = False
    for i, raw in enumerate(raw_rows):
        if i >= len(rows) or not isinstance(raw, dict):
            continue
        snip = raw.get("snippet")
        if snip:
            rows[i][SNIPPET_DISPLAY] = _cell_str(snip)
            has_snippet = True
    if has_snippet and SNIPPET_DISPLAY not in columns_order:
        columns_order.append(SNIPPET_DISPLAY)
    return columns_order, rows


from backend.app.export.column_map import (

    FEE_DATA_START_ROW,

    FEE_HEADER_ROW,

    FEE_SHEET,

    LOCK_DATA_START_ROW,

    LOCK_HEADER_ROW,

    LOCK_SHEET,

    PRODUCT_DATA_ROW,

    PRODUCT_HEADER_ROW,

    PRODUCT_SHEET,

    SHARE_DATA_START_ROW,

    SHARE_HEADER_ROW,

    SHARE_SHEET,
    SUBSCRIPTION_DATA_START_ROW,
    SUBSCRIPTION_HEADER_ROW,
    SUBSCRIPTION_SHEET,
    template_header_for_fee_key,
    template_header_for_subscription_key,
)



PREVIEW_STATUSES = frozenset({"extracted", "exporting", "exported", "export_failed"})





def _cell_str(value: object) -> str | None:

    if value is None:

        return None

    text = str(value).strip()

    return text if text else None





def _read_xlsx_table(

    path: Path,

    sheet_name: str,

    header_row: int,

    data_start_row: int,

    *,

    max_data_rows: int = 100,

) -> tuple[list[str], list[dict[str, str | None]]]:

    wb = load_workbook(path, read_only=True, data_only=True)

    try:

        if sheet_name not in wb.sheetnames:

            return [], []

        ws = wb[sheet_name]

        headers: list[str] = []

        for col in range(1, (ws.max_column or 0) + 1):

            raw = ws.cell(row=header_row, column=col).value

            label = str(raw).strip() if raw is not None else ""

            headers.append(label if label else f"列{col}")



        rows: list[dict[str, str | None]] = []

        for r in range(data_start_row, data_start_row + max_data_rows):

            row_data: dict[str, str | None] = {}

            has_value = False

            for col, header in enumerate(headers, start=1):

                val = _cell_str(ws.cell(row=r, column=col).value)

                row_data[header] = val

                if val:

                    has_value = True

            if not has_value:

                break

            rows.append(row_data)

        return headers, rows

    finally:

        wb.close()





def _product_from_extraction(extraction: dict[str, Any]) -> list[dict[str, str | None]]:

    elements = extraction.get("product_elements") or {}

    rows: list[dict[str, str | None]] = []

    for field, raw in elements.items():

        if isinstance(raw, dict):

            val = raw.get("value")

        else:

            val = raw

        text = _cell_str(val)

        if text:

            rows.append({"field": field, "value": text})

    return rows





def _table_from_extraction_list(

    items: list,

) -> tuple[list[str], list[dict[str, str | None]]]:

    if not items:

        return [], []

    columns_order: list[str] = []

    seen: set[str] = set()

    for row in items:

        if not isinstance(row, dict):

            continue

        for key in row:
            if key in _SKIP_PREVIEW_KEYS:
                continue

            if key not in seen:

                seen.add(key)

                columns_order.append(key)

    rows: list[dict[str, str | None]] = []
    raw_items: list[dict[str, Any]] = []

    for row in items:

        if not isinstance(row, dict):

            continue

        mapped = {
            k: _cell_str(v)
            for k, v in row.items()
            if k not in _SKIP_PREVIEW_KEYS
        }

        if any(mapped.values()):

            rows.append(mapped)
            raw_items.append(row)

    return _append_snippet_column(columns_order, rows, raw_items)





def _fee_from_extraction(extraction: dict[str, Any]) -> tuple[list[str], list[dict[str, str | None]]]:

    fee_list = extraction.get("fee_rates") or []

    if not fee_list:

        return [], []



    columns_order: list[str] = []

    seen: set[str] = set()

    for row in fee_list:

        if not isinstance(row, dict):

            continue

        for key in row:
            if key in _SKIP_PREVIEW_KEYS:
                continue

            label = template_header_for_fee_key(key)

            if label not in seen:

                seen.add(label)

                columns_order.append(label)



    rows: list[dict[str, str | None]] = []
    raw_fee: list[dict[str, Any]] = []

    for row in fee_list:

        if not isinstance(row, dict):

            continue

        mapped: dict[str, str | None] = {}

        for key, val in row.items():
            if key in _SKIP_PREVIEW_KEYS:
                continue

            mapped[template_header_for_fee_key(key)] = _cell_str(val)

        if any(mapped.values()):

            rows.append(mapped)
            raw_fee.append(row)

    return _append_snippet_column(columns_order, rows, raw_fee)


def _subscription_from_extraction(
    extraction: dict[str, Any],
) -> tuple[list[str], list[dict[str, str | None]]]:
    sub_list = extraction.get("subscription_fees") or []
    if not sub_list:
        return [], []

    columns_order: list[str] = []
    seen: set[str] = set()
    for row in sub_list:
        if not isinstance(row, dict):
            continue
        for key in row:
            if key in _SKIP_PREVIEW_KEYS:
                continue
            label = template_header_for_subscription_key(key)
            if label not in seen:
                seen.add(label)
                columns_order.append(label)

    rows: list[dict[str, str | None]] = []
    raw_sub: list[dict[str, Any]] = []
    for row in sub_list:
        if not isinstance(row, dict):
            continue
        mapped: dict[str, str | None] = {}
        for key, val in row.items():
            if key in _SKIP_PREVIEW_KEYS:
                continue
            mapped[template_header_for_subscription_key(key)] = _cell_str(val)
        if any(mapped.values()):
            rows.append(mapped)
            raw_sub.append(row)
    return _append_snippet_column(columns_order, rows, raw_sub)


def build_job_preview(record) -> dict[str, Any]:

    if record.status not in PREVIEW_STATUSES:

        raise ValueError(f"Preview not available for status: {record.status}")



    source = "extraction"

    product_rows: list[dict[str, str | None]] = []

    fee_columns: list[str] = []

    fee_rows: list[dict[str, str | None]] = []

    lock_columns: list[str] = []

    lock_rows: list[dict[str, str | None]] = []

    share_columns: list[str] = []

    share_rows: list[dict[str, str | None]] = []

    subscription_columns: list[str] = []

    subscription_rows: list[dict[str, str | None]] = []

    if record.status == "exported" and getattr(record, "product_xlsx_path", None):

        product_path = (data_dir() / record.product_xlsx_path).resolve()

        if product_path.is_file():

            _, product_data = _read_xlsx_table(

                product_path, PRODUCT_SHEET, PRODUCT_HEADER_ROW, PRODUCT_DATA_ROW, max_data_rows=5

            )

            if product_data:

                wide = product_data[0]

                product_rows = [

                    {"field": k, "value": v} for k, v in wide.items() if v

                ]

            source = "xlsx"



    if record.status == "exported" and getattr(record, "fee_xlsx_path", None):

        fee_path = (data_dir() / record.fee_xlsx_path).resolve()

        if fee_path.is_file():

            fee_columns, fee_rows = _read_xlsx_table(

                fee_path, FEE_SHEET, FEE_HEADER_ROW, FEE_DATA_START_ROW

            )

            source = "xlsx"



    lock_rel = getattr(record, "lock_xlsx_path", None)

    if record.status == "exported" and lock_rel:

        lock_path = (data_dir() / lock_rel).resolve()

        if lock_path.is_file():

            lock_columns, lock_rows = _read_xlsx_table(

                lock_path, LOCK_SHEET, LOCK_HEADER_ROW, LOCK_DATA_START_ROW

            )

            source = "xlsx"



    share_rel = getattr(record, "share_xlsx_path", None)

    if record.status == "exported" and share_rel:

        share_path = (data_dir() / share_rel).resolve()

        if share_path.is_file():

            share_columns, share_rows = _read_xlsx_table(

                share_path, SHARE_SHEET, SHARE_HEADER_ROW, SHARE_DATA_START_ROW

            )

            source = "xlsx"

    sub_rel = getattr(record, "subscription_xlsx_path", None)
    if record.status == "exported" and sub_rel:
        sub_path = (data_dir() / sub_rel).resolve()
        if sub_path.is_file():
            subscription_columns, subscription_rows = _read_xlsx_table(
                sub_path,
                SUBSCRIPTION_SHEET,
                SUBSCRIPTION_HEADER_ROW,
                SUBSCRIPTION_DATA_START_ROW,
            )
            source = "xlsx"

    if record.extraction_result:

        ext = record.extraction_result

        if not product_rows:

            product_rows = _product_from_extraction(ext)

        if not fee_rows:

            fee_columns, fee_rows = _fee_from_extraction(ext)

        if not lock_rows:

            lock_columns, lock_rows = _table_from_extraction_list(

                ext.get("lock_periods") or []

            )

        if not share_rows:

            share_columns, share_rows = _table_from_extraction_list(

                ext.get("share_classes") or []

            )

        if not subscription_rows:
            subscription_columns, subscription_rows = _subscription_from_extraction(ext)

    if (
        not product_rows
        and not fee_rows
        and not lock_rows
        and not share_rows
        and not subscription_rows
    ):
        raise ValueError("No preview data — run extract/export first")

    return {
        "job_id": record.id,
        "source": source,
        "product_rows": product_rows,
        "fee_columns": fee_columns,
        "fee_rows": fee_rows,
        "lock_columns": lock_columns,
        "lock_rows": lock_rows,
        "share_columns": share_columns,
        "share_rows": share_rows,
        "subscription_columns": subscription_columns,
        "subscription_rows": subscription_rows,
    }





def get_job_preview(file_id: uuid.UUID) -> dict[str, Any]:

    from backend.app.db.session import SessionLocal

    from backend.app.models.contract_file import ContractFile



    session = SessionLocal()

    try:

        record = session.get(ContractFile, file_id)

        if record is None:

            raise LookupError(f"contract_file not found: {file_id}")

        return build_job_preview(record)

    finally:

        session.close()


def slice_preview(full: dict[str, Any], section: str) -> dict[str, Any]:
    base = {
        "job_id": full["job_id"],
        "source": full["source"],
        "section": section,
    }
    if section == "product-elements":
        base["product_rows"] = full.get("product_rows") or []
    elif section == "fee-rates":
        base["fee_columns"] = full.get("fee_columns") or []
        base["fee_rows"] = full.get("fee_rows") or []
    elif section == "lock-periods":
        base["lock_columns"] = full.get("lock_columns") or []
        base["lock_rows"] = full.get("lock_rows") or []
    elif section == "share-classes":
        base["share_columns"] = full.get("share_columns") or []
        base["share_rows"] = full.get("share_rows") or []
    elif section == "subscription-fee-rates":
        base["subscription_columns"] = full.get("subscription_columns") or []
        base["subscription_rows"] = full.get("subscription_rows") or []
    else:
        raise ValueError(f"Unknown preview section: {section}")
    return base

