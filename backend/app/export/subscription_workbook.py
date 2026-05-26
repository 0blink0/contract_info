from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.export.column_map import (
    SUBSCRIPTION_DATA_START_ROW,
    SUBSCRIPTION_HEADER_ROW,
    SUBSCRIPTION_SHEET,
    normalize_header,
    template_header_for_subscription_key,
)
from backend.app.export.xlsx_utils import (
    build_header_index,
    clear_data_rows,
    keep_only_sheet,
    write_cell_values,
)


def _subscription_row_values(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, val in row.items():
        if val is None or str(val).strip() == "":
            continue
        template_key = template_header_for_subscription_key(key)
        out[template_key] = val
    return out


def fill_subscription_workbook(
    template_path: Path,
    dest_path: Path,
    subscription_fees: list[dict[str, Any]],
) -> None:
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb[SUBSCRIPTION_SHEET]
    header_index = build_header_index(ws, SUBSCRIPTION_HEADER_ROW)

    display_names: dict[str, str] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=SUBSCRIPTION_HEADER_ROW, column=col).value
        norm = normalize_header(raw)
        if norm and norm not in display_names:
            display_names[norm] = str(raw).strip() if raw else norm

    clear_data_rows(ws, SUBSCRIPTION_DATA_START_ROW)

    for offset, row in enumerate(subscription_fees):
        data = row if isinstance(row, dict) else row.model_dump(by_alias=True)
        values = _subscription_row_values(data)
        mapped: dict[str, Any] = {}
        for k, v in values.items():
            norm = normalize_header(k)
            header_key = display_names.get(norm, k)
            mapped[header_key] = v
            mapped[norm] = v
        write_cell_values(
            ws, SUBSCRIPTION_DATA_START_ROW + offset, header_index, mapped
        )

    keep_only_sheet(wb, SUBSCRIPTION_SHEET)
    wb.save(dest_path)
    wb.close()
