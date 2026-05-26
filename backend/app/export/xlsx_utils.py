from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.export.column_map import normalize_header


def keep_only_sheet(wb: Workbook, sheet_name: str) -> None:
    """Remove all worksheets except *sheet_name* before saving export files."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"sheet {sheet_name!r} not in workbook; have {wb.sheetnames!r}"
        )
    for name in list(wb.sheetnames):
        if name != sheet_name:
            wb.remove(wb[name])


def copy_template(template_path: Path, dest_path: Path) -> None:
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, dest_path)


def build_header_index(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    index: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col).value
        name = normalize_header(cell)
        if not name:
            continue
        index.setdefault(name, []).append(col)
    return index


def clear_data_rows(ws: Worksheet, start_row: int, *, max_col: int | None = None) -> None:
    """Remove template sample rows from *start_row* through the sheet bottom."""
    last_col = max_col or ws.max_column or 1
    end_row = ws.max_row or start_row
    for row in range(start_row, end_row + 1):
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).value = None


def write_cell_values(
    ws: Worksheet,
    row: int,
    header_index: dict[str, list[int]],
    field_to_value: dict[str, Any],
) -> None:
    for field_name, value in field_to_value.items():
        if value is None or value == "":
            continue
        cols = header_index.get(normalize_header(field_name))
        if not cols:
            cols = header_index.get(field_name)
        if not cols:
            continue
        for col in cols:
            ws.cell(row=row, column=col, value=value)
