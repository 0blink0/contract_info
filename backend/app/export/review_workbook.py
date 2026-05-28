"""生成人工核对报告 Excel（路径A字段核对 + 路径B业绩报酬CRM核对）。"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------- 颜色 ----------
_FILL_HEADER = PatternFill("solid", fgColor="2F5496")   # 深蓝表头
_FILL_FULL = PatternFill("solid", fgColor="C6EFCE")     # 绿：可直接填
_FILL_PARTIAL = PatternFill("solid", fgColor="FFEB9C")  # 黄：建议核对
_FILL_MISSING = PatternFill("solid", fgColor="FFC7CE")  # 红：需手录
_FILL_RULE = PatternFill("solid", fgColor="DEEBF7")     # 浅蓝：规则提取
_FILL_LLM = PatternFill("solid", fgColor="EBF3E8")      # 浅绿：LLM提取
_FILL_FIXED = PatternFill("solid", fgColor="F2F2F2")    # 灰：固定值
_FILL_ALT = PatternFill("solid", fgColor="F9F9F9")      # 交替行背景

_FONT_HEADER = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
_FONT_BODY = Font(name="微软雅黑", size=9)
_FONT_EXCERPT = Font(name="微软雅黑", size=8, color="595959")

_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")

_SOURCE_LABEL = {"rule": "规则", "llm": "LLM", "fixed": "固定值", "manual": "人工"}
_CONF_LABEL = {"high": "高", "medium": "中", "low": "低"}
_COVERAGE_LABEL = {"full": "可直接填", "partial": "建议核对", "missing": "需手录"}


def _set_header(ws: Any, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _CENTER


def _coverage_fill(coverage: str) -> PatternFill:
    return {
        "full": _FILL_FULL,
        "partial": _FILL_PARTIAL,
        "missing": _FILL_MISSING,
    }.get(coverage, _FILL_PARTIAL)


def _source_fill(source: str) -> PatternFill:
    return {
        "rule": _FILL_RULE,
        "llm": _FILL_LLM,
        "fixed": _FILL_FIXED,
    }.get(source, _FILL_ALT)


def _write_row(ws: Any, row: int, values: list[Any], fills: list[PatternFill | None] = []) -> None:
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        fill = fills[c - 1] if c - 1 < len(fills) else None
        if fill:
            cell.fill = fill
        is_excerpt = c == len(values) and val and len(str(val)) > 30
        cell.font = _FONT_EXCERPT if is_excerpt else _FONT_BODY
        cell.alignment = _WRAP


def _col_widths(ws: Any, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Sheet 1: 路径 B — CRM 业绩报酬核对 ----------

def _build_path_b_sheet(ws: Any, crm_rows: list[dict], snippet_rows: list[dict]) -> None:
    ws.title = "路径B_CRM核对"
    ws.freeze_panes = "A2"

    # 图例说明
    ws.cell(row=1, column=1, value="【路径B：业绩报酬 / 开放日 CRM 核对表】").font = Font(
        name="微软雅黑", bold=True, size=11, color="2F5496"
    )
    ws.merge_cells("A1:F1")

    headers = ["CRM字段", "建议填写值", "覆盖度", "诊断说明", "合同摘录（参考）"]
    _set_header(ws, 2, headers)

    for r, item in enumerate(crm_rows, start=3):
        cov = item.get("coverage", "missing")
        cov_label = _COVERAGE_LABEL.get(cov, cov)
        fills = [None, _coverage_fill(cov), _coverage_fill(cov), None, None]
        _write_row(
            ws,
            r,
            [
                item.get("crm_field", ""),
                item.get("suggested_value") or "",
                cov_label,
                item.get("diagnostic") or item.get("snippet") or "",
                item.get("snippet") or "",
            ],
            fills,
        )

    # 说明行
    sep = len(crm_rows) + 4
    ws.cell(row=sep, column=1, value="【合同原文摘录（按字段）】").font = Font(
        name="微软雅黑", bold=True, size=10, color="2F5496"
    )
    _set_header(ws, sep + 1, ["字段", "合同摘录"])
    for r, srow in enumerate(snippet_rows, start=sep + 2):
        ws.cell(row=r, column=1, value=srow.get("label") or srow.get("path", "")).font = _FONT_BODY
        ws.cell(row=r, column=2, value=srow.get("text", "")).font = _FONT_EXCERPT
        ws.cell(row=r, column=2).alignment = _WRAP

    _col_widths(ws, [22, 30, 12, 36, 50])
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, 3 + len(crm_rows) + len(snippet_rows) + 10):
        ws.row_dimensions[r].height = 55


# ---------- Sheet 2: 路径 A — 产品要素字段核对 ----------

_PATH_A_DISPLAY_FIELDS: list[str] = [
    "基金全称", "备案编码", "管理人", "托管人", "投资顾问", "外包机构",
    "产品类型（协会）", "基金类型", "基金组织形式", "管理类型",
    "份额结构", "结构类型", "运作方式", "产品存续期",
    "是否量化/对冲基金", "量化策略",
    "首次申购起点", "追加起点", "最低持有类型", "最低持有数量",
    "是否封闭", "封闭方式", "封闭期", "锁定期",
    "是否支持金额赎回", "金额赎回方式",
    "是否支持基金转换", "基金转换方式", "基金转换限制",
    "开放日规则", "最小变动单位",
    "预警线", "止损线",
    "投资目标", "投资范围", "投资策略", "投资限制",
    "风险收益特征", "风险等级", "业绩比较基准",
    "默认分红方式", "投资收益分配",
    "币种", "基金面值",
    "冷静期回访", "双录", "发行机构", "销售机构信息",
    "是否为专户产品",
    "投资经理", "投资经理信息",
    "合伙人类型", "海外基金", "母基金代码", "产品分类", "合同变更方式",
    "自定产品名称",
]

_SKIP_EMPTY_MSG = "合同通常无此字段，系统/运营录入"


def _build_path_a_sheet(ws: Any, product_elements: dict[str, Any]) -> None:
    ws.title = "路径A_字段核对"
    ws.freeze_panes = "A2"

    ws.cell(row=1, column=1, value="【路径A：产品要素字段核对表】").font = Font(
        name="微软雅黑", bold=True, size=11, color="2F5496"
    )
    ws.merge_cells("A1:F1")

    headers = ["字段名", "提取值", "来源", "置信度", "诊断说明", "合同摘录（参考）"]
    _set_header(ws, 2, headers)

    r = 3
    for field in _PATH_A_DISPLAY_FIELDS:
        fv = product_elements.get(field)
        if fv is None:
            value = ""
            source = ""
            conf = ""
            snippet = ""
            diag = _SKIP_EMPTY_MSG
            src_fill = _FILL_ALT
        else:
            if isinstance(fv, dict):
                value = str(fv.get("value") or "")
                source = fv.get("source", "rule")
                conf = fv.get("confidence", "medium")
                snippet = fv.get("snippet") or ""
            else:
                value = str(getattr(fv, "value", "") or "")
                source = getattr(fv, "source", "rule")
                conf = getattr(fv, "confidence", "medium")
                snippet = getattr(fv, "snippet", "") or ""

            src_fill = _source_fill(source)
            source = _SOURCE_LABEL.get(source, source)
            conf_label = _CONF_LABEL.get(conf, conf)
            diag = _build_diag(field, value, conf)
            conf = conf_label

        bg = src_fill if fv else _FILL_MISSING if not value else _FILL_ALT
        fills = [None, bg, None, None, None, None]
        _write_row(ws, r, [field, value, source, conf, diag, snippet[:400] if snippet else ""], fills)
        r += 1

    _col_widths(ws, [22, 30, 8, 6, 30, 60])
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for row in range(3, r + 1):
        ws.row_dimensions[row].height = 55


def _build_diag(field: str, value: str, conf: str) -> str:
    if not value:
        return "未从合同提取，请人工确认"
    long_text_fields = {"投资目标", "投资范围", "投资策略", "投资限制", "风险收益特征"}
    if field in long_text_fields:
        return "长文本字段，建议与原合同核对完整性"
    if conf == "low":
        return "低置信度，建议核对"
    if field in {"产品类型（协会）"}:
        return "协会分类，如为指数增强/对冲类建议核实AMAC登记"
    return ""


# ---------- 主入口 ----------

def build_review_workbook(
    *,
    fund_name: str,
    crm_rows: list[dict],
    snippet_rows: list[dict],
    product_elements: dict[str, Any],
) -> bytes:
    """
    生成人工核对报告，返回 .xlsx 字节流。

    - Sheet 1: 路径B_CRM核对（业绩报酬/开放日）
    - Sheet 2: 路径A_字段核对（产品要素）
    """
    wb = openpyxl.Workbook()
    ws_b = wb.active
    ws_a = wb.create_sheet()

    _build_path_b_sheet(ws_b, crm_rows, snippet_rows)
    _build_path_a_sheet(ws_a, product_elements)

    # 图例 sheet
    ws_legend = wb.create_sheet(title="图例说明")
    _build_legend_sheet(ws_legend, fund_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_legend_sheet(ws: Any, fund_name: str) -> None:
    ws.title = "图例说明"
    ws.cell(row=1, column=1, value=f"核对报告 — {fund_name}").font = Font(
        name="微软雅黑", bold=True, size=12, color="2F5496"
    )
    rows = [
        ("", ""),
        ("颜色说明", ""),
        ("绿色（可直接填）", "提取值置信度高，可直接录入CRM"),
        ("黄色（建议核对）", "有建议值但来自推断，请对照原合同核实"),
        ("红色（需手录）", "未能从合同提取，需人工录入"),
        ("浅蓝（规则提取）", "通过正则/表格规则提取"),
        ("浅绿（LLM提取）", "通过大语言模型推断"),
        ("灰色（固定值）", "系统固定值，无需修改"),
        ("", ""),
        ("来源说明", ""),
        ("规则", "正则/表格直接匹配，高可靠"),
        ("LLM", "语义推断，需核对"),
        ("固定值", "系统预设，无需录入"),
        ("", ""),
        ("注意事项", ""),
        ("1", "本报告仅供参考，最终以合同正文为准"),
        ("2", "提取值较长时已截断，请查阅合同原文"),
        ("3", "业绩报酬字段需与CRM「业绩报酬提取设置」界面逐项对照"),
        ("4", "产品类型（协会）如为指数增强/量化对冲类，建议核实AMAC登记"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(name="微软雅黑", bold=bool(v == ""), size=9)
        ws.cell(row=i, column=2, value=v).font = Font(name="微软雅黑", size=9)
    _col_widths(ws, [28, 50])
