from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.export.column_map import (
    FEE_DATA_START_ROW,
    FEE_HEADER_ROW,
    FEE_SHEET,
    normalize_header,
    template_header_for_fee_key,
)
from backend.app.export.xlsx_utils import build_header_index, write_cell_values


def _fee_row_values(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, val in row.items():
        if val is None or str(val).strip() == "":
            continue
        template_key = template_header_for_fee_key(key)
        out[template_key] = val
    return out


def fill_fee_workbook(
    template_path: Path,
    dest_path: Path,
    fee_rates: list[dict[str, Any]],
) -> None:
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb[FEE_SHEET]
    header_index = build_header_index(ws, FEE_HEADER_ROW)

    # Map normalized header -> template display name from row
    display_names: dict[str, str] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=FEE_HEADER_ROW, column=col).value
        norm = normalize_header(raw)
        if norm and norm not in display_names:
            display_names[norm] = str(raw).strip() if raw else norm

    for offset, row in enumerate(fee_rates):
        data = row if isinstance(row, dict) else row.model_dump(by_alias=True)
        values = _fee_row_values(data)
        # Use first cell text for header match (may include 【必填】)
        mapped: dict[str, Any] = {}
        for k, v in values.items():
            norm = normalize_header(k)
            header_key = display_names.get(norm, k)
            mapped[header_key] = v
            mapped[norm] = v
        write_cell_values(ws, FEE_DATA_START_ROW + offset, header_index, mapped)

    wb.save(dest_path)
    wb.close()
