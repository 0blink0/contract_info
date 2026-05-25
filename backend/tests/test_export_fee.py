import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.config import templates_dir
from backend.app.export.column_map import FEE_DATA_START_ROW, FEE_HEADER_ROW, FEE_SHEET, normalize_header
from backend.app.export.fee_workbook import fill_fee_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "sample_extraction.json"


@pytest.fixture
def sample_extraction():
    return json.loads(FIXTURE.read_text(encoding="utf-8"))


def test_fill_fee_workbook(tmp_path, sample_extraction):
    dest = tmp_path / "fee.xlsx"
    fill_fee_workbook(
        templates_dir() / "产品运营费率导入模板-1.xlsx",
        dest,
        sample_extraction["fee_rates"],
    )
    wb = load_workbook(dest, read_only=True)
    ws = wb[FEE_SHEET]
    type_col = None
    rate_col = None
    for c in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(FEE_HEADER_ROW, c).value)
        if h == "运营费类型":
            type_col = c
        if "费率" in h and "%" in h:
            rate_col = c
    assert type_col and rate_col
    types = {
        ws.cell(FEE_DATA_START_ROW + i, type_col).value for i in range(2)
    }
    assert "管理费" in types
    assert "托管费" in types
    wb.close()
