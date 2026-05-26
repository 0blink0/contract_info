from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.export.column_map import (
    FEE_DATA_START_ROW,
    FEE_HEADER_ROW,
    FEE_SHEET,
    LOCK_DATA_START_ROW,
    LOCK_HEADER_ROW,
    LOCK_SHEET,
    PRODUCT_DATA_ROW,
    PRODUCT_HEADER_ROW,
    PRODUCT_SHEET,
    SHARE_HEADER_ROW,
    SHARE_SHEET,
    SUBSCRIPTION_DATA_START_ROW,
    SUBSCRIPTION_HEADER_ROW,
    SUBSCRIPTION_SHEET,
    normalize_header,
)
from backend.tests.golden.helpers.normalize import (
    contains_normalized,
    empty_equiv,
    normalize_cell,
    normalize_party_name,
    normalize_pct,
)

CRITICAL_FROM_CONTRACT = frozenset(
    {
        "基金全称",
        "基金简称",
        "管理人",
        "托管人",
        "外包机构",
        "投资顾问",
        "风险等级",
        "投资经理",
        "锁定期",
        "开放日规则",
        "预警线",
        "止损线",
    }
)

A_CLASS_EMPTY_OK = frozenset(
    {
        "基金代码",
        "成立日期",
        "备案日期",
    }
)

SKIP_GOLDEN_PARTY = frozenset({"管理人", "托管人", "外包机构", "投资顾问"})


def _header_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_header(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def find_product_row(wb_path: Path, fund_full_name: str) -> int:
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[PRODUCT_SHEET]
    headers = _header_map(ws, PRODUCT_HEADER_ROW)
    name_col = headers.get(normalize_header("基金全称"))
    if not name_col:
        wb.close()
        raise AssertionError(f"基金全称 column not found in {wb_path}")
    target = normalize_cell(fund_full_name)
    for row in range(PRODUCT_DATA_ROW, ws.max_row + 1):
        val = normalize_cell(ws.cell(row, name_col).value)
        if val == target or target in val or val in target:
            wb.close()
            return row
    wb.close()
    raise AssertionError(f"Fund row not found for {fund_full_name!r} in {wb_path}")


def read_product_field(wb_path: Path, row: int, header_name: str) -> str | None:
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[PRODUCT_SHEET]
    headers = _header_map(ws, PRODUCT_HEADER_ROW)
    col = headers.get(normalize_header(header_name))
    if not col:
        wb.close()
        return None
    val = ws.cell(row, col).value
    wb.close()
    return normalize_cell(val) if val is not None else None


def assert_critical_product(
    export_path: Path,
    expected: dict[str, Any],
    *,
    skip_party_from_golden: bool = True,
) -> None:
    fund = expected.get("基金全称") or ""
    row = find_product_row(export_path, fund)
    for field in CRITICAL_FROM_CONTRACT:
        if field not in expected:
            continue
        exp_val = expected[field]
        actual = read_product_field(export_path, row, field)
        if field in ("管理人", "托管人", "外包机构", "投资顾问"):
            if empty_equiv(exp_val, actual):
                continue
            assert normalize_party_name(actual) == normalize_party_name(
                exp_val
            ), f"{field}: expected {exp_val!r}, got {actual!r}"
        elif field in ("预警线", "止损线", "锁定期"):
            assert empty_equiv(actual, exp_val) or contains_normalized(
                actual, exp_val
            ), f"{field}: expected {exp_val!r}, got {actual!r}"
        elif field == "开放日规则":
            assert contains_normalized(
                actual, exp_val
            ), f"{field}: expected substring {exp_val!r}, got {actual!r}"
        elif field == "投资经理":
            for part in str(exp_val).replace("、", ",").split(","):
                part = part.strip()
                if part:
                    assert contains_normalized(
                        actual, part
                    ), f"{field}: missing {part!r} in {actual!r}"
        else:
            assert empty_equiv(actual, exp_val) or normalize_cell(
                actual
            ) == normalize_cell(exp_val), (
                f"{field}: expected {exp_val!r}, got {actual!r}"
            )
    _ = skip_party_from_golden


def assert_fee_types_present(
    export_path: Path,
    fund_name: str,
    expected_by_type: dict[str, Any],
) -> None:
    wb = load_workbook(export_path, read_only=True, data_only=True)
    ws = wb[FEE_SHEET]
    headers = _header_map(ws, FEE_HEADER_ROW)
    type_col = headers.get(normalize_header("运营费类型"))
    fund_col = headers.get(normalize_header("基金名称"))
    rate_col = headers.get(normalize_header("费率（单位：%/年）")) or headers.get(
        normalize_header("费率（%/年）")
    )
    if not type_col or not rate_col:
        wb.close()
        raise AssertionError("Fee sheet missing 运营费类型 or 费率 column")

    found: dict[str, list[str]] = {}
    target_fund = normalize_cell(fund_name)
    fund_prefix = target_fund[:12] if len(target_fund) >= 8 else target_fund

    def _same_fund(fname: str) -> bool:
        if not fname:
            return True
        if target_fund in fname or fname in target_fund:
            return True
        return bool(fund_prefix and fname.startswith(fund_prefix))

    for row in range(FEE_DATA_START_ROW, ws.max_row + 1):
        fname = normalize_cell(ws.cell(row, fund_col).value) if fund_col else ""
        if fund_col and fname and not _same_fund(fname):
            continue
        fee_type = normalize_cell(ws.cell(row, type_col).value)
        if not fee_type:
            continue
        rate = normalize_pct(ws.cell(row, rate_col).value)
        found.setdefault(fee_type, []).append(rate)

    wb.close()
    for fee_type, spec in expected_by_type.items():
        exp_rate = spec.get("rate_annual_pct") if isinstance(spec, dict) else spec
        rates = found.get(fee_type) or []
        assert rates, f"Missing fee type {fee_type!r} in export"
        exp = normalize_pct(exp_rate)
        assert any(normalize_pct(r) == exp for r in rates), (
            f"{fee_type}: expected rate {exp_rate!r} among rows, got {rates!r}"
        )


def assert_export_structure(paths: dict[str, Path]) -> None:
    for key, path in paths.items():
        assert path.is_file(), f"Missing export file {key}: {path}"
    wb = load_workbook(paths["product"], read_only=True)
    assert PRODUCT_SHEET in wb.sheetnames
    wb.close()
    wb = load_workbook(paths["fee"], read_only=True)
    assert FEE_SHEET in wb.sheetnames
    wb.close()
    wb = load_workbook(paths["lock"], read_only=True)
    assert LOCK_SHEET in wb.sheetnames
    wb.close()
    wb = load_workbook(paths["share"], read_only=True)
    assert SHARE_SHEET in wb.sheetnames
    wb.close()
    if "subscription" in paths:
        wb = load_workbook(paths["subscription"], read_only=True)
        assert SUBSCRIPTION_SHEET in wb.sheetnames
        wb.close()


def assert_subscription_rates(
    export_path: Path,
    expected_by_share: dict[str, Any],
) -> None:
    wb = load_workbook(export_path, read_only=True, data_only=True)
    ws = wb[SUBSCRIPTION_SHEET]
    headers = _header_map(ws, SUBSCRIPTION_HEADER_ROW)
    name_col = headers.get(normalize_header("基金名称"))
    type_col = headers.get(normalize_header("申赎费类型"))
    rate_col = headers.get(normalize_header("费率"))
    if not name_col or not type_col or not rate_col:
        wb.close()
        raise AssertionError("Subscription sheet missing required columns")

    found: dict[str, dict[str, str]] = {}
    for row in range(SUBSCRIPTION_DATA_START_ROW, ws.max_row + 1):
        fname = normalize_cell(ws.cell(row, name_col).value) or ""
        fee_type = normalize_cell(ws.cell(row, type_col).value) or ""
        if not fname or not fee_type:
            continue
        rate = normalize_pct(ws.cell(row, rate_col).value)
        for letter in expected_by_share:
            if letter.upper() in fname.upper():
                found.setdefault(letter.upper(), {})[fee_type] = rate
                break

    wb.close()
    for letter, rates in expected_by_share.items():
        for fee_type, exp_rate in rates.items():
            got = found.get(letter.upper(), {}).get(fee_type)
            assert got is not None, f"Missing {letter} {fee_type}"
            assert normalize_pct(got) == normalize_pct(
                exp_rate
            ), f"{letter} {fee_type}: expected {exp_rate!r}, got {got!r}"


def assert_lock_sheet(
    export_lock_path: Path,
    fund_name: str,
    *,
    expect_rows: int | None,
) -> None:
    wb = load_workbook(export_lock_path, read_only=True, data_only=True)
    ws = wb[LOCK_SHEET]
    headers = _header_map(ws, LOCK_HEADER_ROW)
    name_col = headers.get(normalize_header("产品名称"))
    count = 0
    target = normalize_cell(fund_name)
    for row in range(LOCK_DATA_START_ROW, ws.max_row + 1):
        if name_col:
            val = normalize_cell(ws.cell(row, name_col).value)
            if val and target not in val and val not in target:
                continue
        if any(ws.cell(row, c).value for c in range(1, ws.max_column + 1)):
            count += 1
    wb.close()
    if expect_rows is not None:
        assert count == expect_rows, (
            f"Lock sheet rows for {fund_name}: expected {expect_rows}, got {count}"
        )


def assert_share_sheet_extended(export_share_path: Path) -> None:
    wb = load_workbook(export_share_path, read_only=True)
    assert SHARE_SHEET in wb.sheetnames
    wb.close()
