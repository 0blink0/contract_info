import json
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.export.column_map import (
    SUBSCRIPTION_DATA_START_ROW,
    SUBSCRIPTION_HEADER_ROW,
    SUBSCRIPTION_SHEET,
    normalize_header,
)
from backend.app.export.pipeline import export_xlsx

FIXTURE = Path(__file__).parent / "fixtures" / "sample_extraction.json"


def test_subscription_workbook_sheet_name(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "backend.app.export.pipeline.exports_dir",
        lambda: tmp_path / "exports",
    )
    extraction = json.loads(FIXTURE.read_text(encoding="utf-8"))
    fid = uuid.uuid4()
    *_, sub_rel, _ = export_xlsx(extraction, fid)
    sub_path = tmp_path / "exports" / str(fid) / Path(sub_rel).name
    wb = load_workbook(sub_path, read_only=True)
    assert SUBSCRIPTION_SHEET in wb.sheetnames
    ws = wb[SUBSCRIPTION_SHEET]
    headers = {
        normalize_header(ws.cell(SUBSCRIPTION_HEADER_ROW, c).value)
        for c in range(1, 15)
    }
    assert "基金名称" in headers
    assert "申赎费类型" in headers
    val = ws.cell(SUBSCRIPTION_DATA_START_ROW, 1).value
    assert val and "正仁" in str(val)
    wb.close()


@pytest.mark.parametrize("docx_key", [
    "石云中证1000资产进取一号私募证券投资基金-基金合同(1).docx",
    "石云福禄1000指数增强一号私募证券投资基金(1).docx",
])
def test_subscription_export_from_contract(docx_key, tmp_path, monkeypatch):
    from backend.app.extract.pipeline import extract_document_sync
    from backend.app.parse import parse_docx
    from backend.app.parse.schemas import document_to_dict

    docx = Path(__file__).resolve().parents[2] / "example" / docx_key
    if not docx.is_file():
        pytest.skip(f"missing {docx}")

    class LlmOff:
        available = False
        model_name = ""

    monkeypatch.setattr(
        "backend.app.export.pipeline.exports_dir",
        lambda: tmp_path / "exports",
    )
    doc = document_to_dict(parse_docx(str(docx)))
    result, _, _path_b = extract_document_sync(doc, llm_client=LlmOff())  # type: ignore[arg-type]
    fid = uuid.uuid4()
    *_, sub_rel, _ = export_xlsx(result.model_dump(mode="json"), fid)
    sub_path = tmp_path / "exports" / str(fid) / Path(sub_rel).name
    assert sub_path.is_file()
    wb = load_workbook(sub_path, read_only=True)
    ws = wb[SUBSCRIPTION_SHEET]
    names: list[str] = []
    for r in range(SUBSCRIPTION_DATA_START_ROW, SUBSCRIPTION_DATA_START_ROW + 30):
        val = ws.cell(r, 1).value
        if not val:
            break
        names.append(str(val))
    wb.close()
    expected_rows = len(result.subscription_fees)
    assert len(names) == expected_rows
    assert len(names) >= 8
    if "福禄" not in docx_key:
        assert not any("福禄" in n for n in names)
    else:
        assert not any("中证" in n for n in names)
