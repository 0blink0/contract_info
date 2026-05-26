"""Each export xlsx retains only its business sheet (no extra template tabs)."""

import json
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.export.column_map import (
    FEE_SHEET,
    LOCK_SHEET,
    PRODUCT_SHEET,
    SHARE_SHEET,
    SUBSCRIPTION_SHEET,
)
from backend.app.export.pipeline import export_xlsx
from backend.app.config import templates_dir
from backend.app.export.product_workbook import fill_product_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "sample_extraction.json"


@pytest.fixture
def sample_extraction():
    return json.loads(FIXTURE.read_text(encoding="utf-8"))


def test_product_workbook_single_sheet(tmp_path, sample_extraction):
    dest = tmp_path / "product.xlsx"
    fill_product_workbook(
        templates_dir() / "产品要素-2.xlsx",
        dest,
        sample_extraction["product_elements"],
    )
    wb = load_workbook(dest, read_only=True)
    assert wb.sheetnames == [PRODUCT_SHEET]
    wb.close()


def test_export_pipeline_single_sheet_per_file(tmp_path, monkeypatch, sample_extraction):
    monkeypatch.setattr(
        "backend.app.export.pipeline.exports_dir",
        lambda: tmp_path / "exports",
    )
    fid = uuid.uuid4()
    product, fee, lock, share, sub, _ = export_xlsx(sample_extraction, fid)
    base = tmp_path / "exports" / str(fid)
    cases = [
        (base / Path(product).name, PRODUCT_SHEET),
        (base / Path(fee).name, FEE_SHEET),
        (base / Path(lock).name, LOCK_SHEET),
        (base / Path(share).name, SHARE_SHEET),
        (base / Path(sub).name, SUBSCRIPTION_SHEET),
    ]
    for path, sheet in cases:
        wb = load_workbook(path, read_only=True)
        assert wb.sheetnames == [sheet], f"{path.name}: {wb.sheetnames}"
        wb.close()
