import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.config import PROJECT_ROOT, templates_dir
from backend.app.export.column_map import PRODUCT_DATA_ROW, PRODUCT_HEADER_ROW, PRODUCT_SHEET
from backend.app.export.product_workbook import fill_product_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "sample_extraction.json"


@pytest.fixture
def sample_extraction():
    return json.loads(FIXTURE.read_text(encoding="utf-8"))


def test_fill_product_workbook(tmp_path, sample_extraction):
    dest = tmp_path / "out.xlsx"
    fill_product_workbook(
        templates_dir() / "产品要素-2.xlsx",
        dest,
        sample_extraction["product_elements"],
    )
    wb = load_workbook(dest, read_only=True)
    ws = wb[PRODUCT_SHEET]
    headers = [
        ws.cell(PRODUCT_HEADER_ROW, c).value for c in range(1, ws.max_column + 1) if ws.cell(PRODUCT_HEADER_ROW, c).value
    ]
    assert "基金全称" in headers
    # find column for 基金全称
    col = next(
        i
        for i in range(1, ws.max_column + 1)
        if ws.cell(PRODUCT_HEADER_ROW, i).value == "基金全称"
    )
    assert ws.cell(PRODUCT_DATA_ROW, col).value == "正仁1号私募证券投资基金"
    wb.close()


def test_templates_committed():
    assert (templates_dir() / "产品要素-2.xlsx").is_file()
