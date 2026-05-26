import json
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.config import PROJECT_ROOT
from backend.app.export.column_map import PRODUCT_DATA_ROW, PRODUCT_SHEET
from backend.app.export.pipeline import export_xlsx

FIXTURE = Path(__file__).parent / "fixtures" / "sample_extraction.json"


def test_export_xlsx_paths(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "backend.app.export.pipeline.exports_dir",
        lambda: tmp_path / "exports",
    )
    extraction = json.loads(FIXTURE.read_text(encoding="utf-8"))
    fid = uuid.uuid4()
    product_rel, fee_rel, lock_rel, share_rel, sub_rel, warnings = export_xlsx(
        extraction, fid
    )
    product_abs = tmp_path / "exports" / str(fid) / "product_elements.xlsx"
    lock_abs = tmp_path / "exports" / str(fid) / "lock_periods.xlsx"
    share_abs = tmp_path / "exports" / str(fid) / "share_classes.xlsx"
    assert product_abs.is_file()
    assert lock_abs.is_file()
    assert share_abs.is_file()
    assert lock_rel.endswith("lock_periods.xlsx")
    assert share_rel.endswith("share_classes.xlsx")
    sub_abs = tmp_path / "exports" / str(fid) / "subscription_fee_rates.xlsx"
    assert sub_abs.is_file()
    assert sub_rel.endswith("subscription_fee_rates.xlsx")
    wb = load_workbook(product_abs, read_only=True)
    ws = wb[PRODUCT_SHEET]
    assert ws.cell(PRODUCT_DATA_ROW, 2).value  # 基金全称 col 2 typical
    wb.close()
